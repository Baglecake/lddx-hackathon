#!/usr/bin/env python3
"""
Batch runner for LDDx pipeline evaluation.

Randomly samples cases from Open-XDDx.xlsx and runs them through the
pipeline, saving results as JSON for downstream evaluation.

Usage:
    python batch_run.py --n 20 --mode quick --backend mlx --model mlx-community/gemma-4-26b-a4b-it-4bit
    python batch_run.py --n 20 --mode full --backend ollama --model qwen2.5:32b-instruct-q8_0
"""

import argparse
import ast
import json
import os
import random
import sys
import time
from datetime import datetime

import openpyxl

# Pipeline imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pipeline', 'Modules'))

from ddx_core import ModelConfig, OllamaModelManager, DynamicAgentGenerator
from ddx_runner import DDxSystem
from ddx_sliding_context import TranscriptManager


def load_cases(xlsx_path: str) -> list[dict]:
    """Load all cases from the Open-XDDx spreadsheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Sheet1']

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    cases = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        case = dict(zip(headers, row))
        if not case.get('patient_info'):
            continue

        # Parse the interpretation dict (ground truth)
        interp = case.get('interpretation', '')
        if isinstance(interp, str) and interp.startswith('{'):
            try:
                case['ground_truth'] = ast.literal_eval(interp)
            except (ValueError, SyntaxError):
                case['ground_truth'] = {}
        else:
            case['ground_truth'] = {}

        cases.append(case)

    wb.close()
    return cases


def initialize_system(backend: str, model: str, url: str) -> DDxSystem:
    """Initialize the DDx system with the specified backend and model."""
    system = DDxSystem()

    configs = {
        'conservative_model': ModelConfig(
            name='Conservative', model_name=model,
            temperature=0.3, top_p=0.7, max_tokens=1024, role='conservative',
        ),
        'innovative_model': ModelConfig(
            name='Innovative', model_name=model,
            temperature=0.8, top_p=0.95, max_tokens=1024, role='innovative',
        ),
    }

    backend_kwargs = {}
    if backend == 'ollama':
        backend_kwargs['base_url'] = url

    system.model_manager = OllamaModelManager(
        configs, backend_type=backend, **backend_kwargs
    )

    if not system.model_manager.initialize():
        print("ERROR: Failed to initialize backend")
        sys.exit(1)

    for model_id in system.model_manager.get_available_models():
        system.model_manager.load_model(model_id)

    system.transcript = TranscriptManager()
    system.agent_generator = DynamicAgentGenerator(
        system.model_manager, system.transcript
    )

    return system


def run_batch(system: DDxSystem, cases: list[dict], mode: str,
              max_specialists: int = 5,
              output_dir: str = 'exports', run_metadata: dict = None) -> list[dict]:
    """Run the pipeline on a batch of cases, saving each as its own JSON."""
    results = []
    os.makedirs(output_dir, exist_ok=True)

    for i, case in enumerate(cases, 1):
        case_idx = case.get('Index', i)
        patient_info = case['patient_info']
        ground_truth = case.get('ground_truth', {})
        specialty = case.get('specialty', '')

        print(f"\n{'='*70}")
        print(f"CASE {i}/{len(cases)} (Index: {case_idx})")
        print(f"Specialty: {specialty}")
        print(f"Ground truth diagnoses: {list(ground_truth.keys())}")
        print(f"{'='*70}")
        print(f"Patient info: {patient_info[:150]}...")

        start_time = time.time()

        # Analyze case
        analysis = system.analyze_case(
            patient_info, f"case_{case_idx}", max_specialists=max_specialists
        )

        if not analysis.get('success'):
            print(f"  FAILED: {analysis.get('error')}")
            case_result = {
                'case_index': case_idx,
                'error': analysis.get('error'),
                'ground_truth': ground_truth,
                'specialty': specialty,
            }
            results.append(case_result)
            # Save even failed cases
            case_file = os.path.join(output_dir, f"case_{case_idx}.json")
            with open(case_file, 'w') as f:
                json.dump(case_result, f, indent=2, default=str)
            print(f"  Saved → {case_file}")
            continue

        # Run pipeline
        if mode == 'full':
            result = system.run_full_diagnosis()
        else:
            result = system.run_quick_diagnosis()

        duration = time.time() - start_time

        # Extract final diagnoses
        final_dx = []
        voting = result.get('voting_result')
        if voting and voting.get('ranked'):
            final_dx = [d[0] for d in voting['ranked'][:6]]
        elif result.get('final_diagnoses'):
            final_dx = result['final_diagnoses'][:6]

        # Build case result
        case_result = {
            'case_index': case_idx,
            'patient_info': patient_info,
            'specialty': specialty,
            'ground_truth': ground_truth,
            'ground_truth_diagnoses': list(ground_truth.keys()),
            'pipeline_diagnoses': final_dx,
            'specialists': analysis.get('specialists', []),
            'rounds': {
                rnd: {
                    'response_count': len(rdata.get('responses', [])),
                    'duration': rdata.get('duration', 0),
                }
                for rnd, rdata in result.get('rounds', {}).items()
            },
            'credibility_scores': result.get('credibility_scores', {}),
            'voting_result': result.get('voting_result'),
            'total_duration': duration,
            'full_results': result,
            'run_metadata': run_metadata,
        }

        results.append(case_result)

        # Save individual case JSON
        case_file = os.path.join(output_dir, f"case_{case_idx}.json")
        with open(case_file, 'w') as f:
            json.dump(case_result, f, indent=2, default=str)

        print(f"\n  Duration: {duration:.1f}s")
        print(f"  Pipeline diagnoses: {final_dx[:3]}")
        print(f"  Ground truth: {list(ground_truth.keys())[:3]}")
        print(f"  Saved → {case_file}")

    return results


def stratified_sample(cases: list[dict], specialties: dict[str, int],
                       seed: int = 42) -> list[dict]:
    """
    Sample cases stratified by specialty.

    Args:
        cases: All available cases
        specialties: Dict mapping specialty name to number of cases to sample
                     e.g. {'Cardiovascular disease': 4, 'Nervous system disease': 4}
        seed: Random seed
    """
    random.seed(seed)
    sampled = []

    for spec, n in specialties.items():
        pool = [c for c in cases if c.get('specialty', '').lower() == spec.lower()]
        if not pool:
            # Try partial match
            pool = [c for c in cases if spec.lower() in c.get('specialty', '').lower()]
        if len(pool) < n:
            print(f"  Warning: only {len(pool)} cases for '{spec}', requested {n}")
            n = len(pool)
        chosen = random.sample(pool, n)
        sampled.extend(chosen)
        print(f"  {spec}: sampled {len(chosen)}/{len(pool)} available")

    return sampled


def main():
    parser = argparse.ArgumentParser(description="LDDx Batch Runner")
    parser.add_argument('--n', type=int, default=4,
                        help='Cases per specialty for stratified sampling (default: 4)')
    parser.add_argument('--mode', choices=['quick', 'full'], default='full',
                        help='Pipeline mode (default: full)')
    parser.add_argument('--backend', choices=['ollama', 'mlx'], default='mlx',
                        help='Inference backend (default: mlx)')
    parser.add_argument('--model', type=str,
                        default='mlx-community/gemma-4-26b-a4b-it-4bit',
                        help='Model name')
    parser.add_argument('--url', type=str, default='http://localhost:11434',
                        help='Ollama URL (only used with ollama backend)')
    parser.add_argument('--seed', type=int, default=42,
                        help='Random seed for reproducibility (default: 42)')
    parser.add_argument('--specialists', type=int, default=5,
                        help='Max specialists per case (default: 5)')
    parser.add_argument('--output', type=str, default=None,
                        help='Output JSON path (auto-generated if not set)')
    parser.add_argument('--data', type=str, default='data/Open-XDDx.xlsx',
                        help='Path to dataset')
    parser.add_argument('--specialties', type=str, nargs='+',
                        default=[
                            'Cardiovascular disease',
                            'Circulatory system disease',
                            'Digestive system disease',
                            'Endocrine disorder',
                            'Nervous system disease',
                        ],
                        help='Specialties to sample from')
    parser.add_argument('--random-sample', action='store_true',
                        help='Use random sampling instead of stratified')

    args = parser.parse_args()

    # Load cases
    print(f"Loading cases from {args.data}...")
    all_cases = load_cases(args.data)
    print(f"Loaded {len(all_cases)} cases")

    # Sample
    if args.random_sample:
        random.seed(args.seed)
        sample_size = min(args.n, len(all_cases))
        sampled = random.sample(all_cases, sample_size)
        print(f"Randomly sampled {sample_size} cases (seed={args.seed})")
    else:
        spec_dict = {s: args.n for s in args.specialties}
        print(f"\nStratified sampling ({args.n} per specialty):")
        sampled = stratified_sample(all_cases, spec_dict, args.seed)
        sample_size = len(sampled)
        print(f"Total: {sample_size} cases")

    # Initialize
    print(f"\nInitializing {args.backend} backend with {args.model}...")
    system = initialize_system(args.backend, args.model, args.url)

    # Output directory — one JSON per case
    output_dir = args.output or (
        f"exports/batch_{args.backend}_{args.mode}_{sample_size}cases_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )

    run_metadata = {
        'backend': args.backend,
        'model': args.model,
        'mode': args.mode,
        'n_cases': sample_size,
        'seed': args.seed,
        'max_specialists': args.specialists,
        'timestamp': datetime.now().isoformat(),
    }

    # Run
    print(f"\nRunning {args.mode} pipeline on {sample_size} cases...")
    print(f"Each case saves to: {output_dir}/case_<index>.json")
    batch_start = time.time()
    results = run_batch(system, sampled, args.mode, args.specialists,
                        output_dir=output_dir, run_metadata=run_metadata)
    batch_duration = time.time() - batch_start

    # Final summary
    successful = [r for r in results if 'error' not in r]
    failed = [r for r in results if 'error' in r]
    avg_time = sum(r.get('total_duration', 0) for r in successful) / len(successful) if successful else 0

    print(f"\n{'='*70}")
    print(f"BATCH COMPLETE")
    print(f"{'='*70}")
    print(f"  Cases: {len(results)} ({len(successful)} successful, {len(failed)} failed)")
    print(f"  Total time: {batch_duration:.1f}s")
    print(f"  Avg time per case: {avg_time:.1f}s")
    print(f"  Backend: {args.backend} / {args.model}")
    print(f"  Mode: {args.mode}")
    print(f"  Output: {output_dir}/")

    # Save batch summary alongside individual case files
    summary_path = os.path.join(output_dir, '_batch_summary.json')
    summary = {
        **run_metadata,
        'total_duration': batch_duration,
        'avg_duration': avg_time,
        'successful': len(successful),
        'failed': len(failed),
        'status': 'completed',
        'case_files': [f"case_{r['case_index']}.json" for r in results],
    }
    with open(summary_path, 'w') as f:
        json.dump(summary, f, indent=2, default=str)
    print(f"  Summary: {summary_path}")

    print(f"\n  Results saved to: {output_path}")


if __name__ == '__main__':
    main()
